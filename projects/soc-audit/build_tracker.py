"""Build SOC 2026 Audit Tracker.xlsx from evidence-log data."""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

WB_PATH = r"c:\Users\rsnide\OneDrive - Revolution Group\Revolution Group - SOC Audit - 2026 - SOC Audit\SOC 2026 Audit Tracker.xlsx"

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
RED_FILL    = PatternFill("solid", fgColor="FFCCCC")
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
GREEN_FILL  = PatternFill("solid", fgColor="D9EAD3")
BLUE_FILL   = PatternFill("solid", fgColor="CFE2F3")
GREY_FILL   = PatternFill("solid", fgColor="EEEEEE")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")

HDR_FILL    = PatternFill("solid", fgColor="1F4E79")   # dark navy
SUB_FILL    = PatternFill("solid", fgColor="2E75B6")   # section header
HDR_FONT    = Font(bold=True, color="FFFFFF", size=10)
SUB_FONT    = Font(bold=True, color="FFFFFF", size=10)
NORMAL_FONT = Font(size=10)
BOLD_FONT   = Font(bold=True, size=10)

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

STATUS_FILL = {
    "collected":    GREEN_FILL,
    "in-progress":  YELLOW_FILL,
    "missing":      RED_FILL,
    "remote":       BLUE_FILL,
    "tbd":          GREY_FILL,
    "n/a":          GREY_FILL,
}

# ---------------------------------------------------------------------------
# Data — all 116 controls
# ---------------------------------------------------------------------------
# Columns: matrix_num, matrix_name, control, description, type, status,
#          assigned_to, evidence_location, notes

CONTROLS = [
    # Matrix 1 — Control Environment
    (1,"Control Environment","1.1","Employee handbook","provide","collected","","Deliverables\\Employee-Handbookv4.pdf",""),
    (1,"Control Environment","1.2","Full staff list (name, hire date, term date — all during review period)","provide","collected","","Deliverables\\1.2,1.4,1.22 Staffing and Performance Reviews.xlsx",""),
    (1,"Control Environment","1.3","Background checks — sample employees","remote","remote","","","Pulled from 1.2 sample"),
    (1,"Control Environment","1.4","Contract employees (1099) hired since June 1 2025 — names and hire dates","provide","collected","","Deliverables\\1.2,1.4,1.22 Staffing and Performance Reviews.xlsx",""),
    (1,"Control Environment","1.5","Background checks — contract employees","remote","remote","","","Pulled from 1.4 sample"),
    (1,"Control Environment","1.6","Certificate of insurance (COI)","provide","collected","","Deliverables\\COI - Revolution Group 2026-2027.pdf",""),
    (1,"Control Environment","1.7","2 written job descriptions","provide","collected","","Deliverables\\1.7;1.9-TSD_Help Desk Consultant.docx + 1.7;1.9-TSD_IT Account Manager Consultant.docx + 1.7;1.9-TSD_Technical Services Tier II.docx",""),
    (1,"Control Environment","1.8","Candidate evaluation process","tbd","tbd","","","Confirm with Audit Manager"),
    (1,"Control Environment","1.9","2 job descriptions noting duties to interact/monitor third-party vendors","provide","collected","","Deliverables\\1.7;1.9-TSD_Help Desk Consultant.docx + 1.7;1.9-TSD_IT Account Manager Consultant.docx + 1.7;1.9-TSD_Technical Services Tier II.docx + 1.9 Note on Third-Party interactions.docx",""),
    (1,"Control Environment","1.10","Training and development program","tbd","tbd","","","Confirm with Audit Manager"),
    (1,"Control Environment","1.11","Continuing education encouragement","tbd","tbd","","","Confirm with Audit Manager"),
    (1,"Control Environment","1.12","Professional development expense reimbursement","tbd","tbd","","","Confirm with Audit Manager"),
    (1,"Control Environment","1.13","Board of directors oversight","provide","collected","","Deliverables\\1.13-Board of Directors.docx",""),
    (1,"Control Environment","1.14","BOD meeting dates since June 1 2025","provide","collected","","Deliverables\\1.14 - BOD meeting agendas from most recent meeting.docx",""),
    (1,"Control Environment","1.15","InfoSec importance communicated to employees","tbd","tbd","","","Confirm with Audit Manager"),
    (1,"Control Environment","1.16","Corporate mission statement","provide","collected","","Deliverables\\1.16-Mission Statement and Core Values.docx",""),
    (1,"Control Environment","1.17","Trade shows attended, publications used, or industry association memberships","provide","collected","","Deliverables\\1.17-List of Tradeshows and memberships.docx",""),
    (1,"Control Environment","1.18","Management meeting schedule/cadence","tbd","tbd","","","Confirm with Audit Manager"),
    (1,"Control Environment","1.19","Updated organizational chart","provide","collected","","Deliverables\\1.19-Organizational Chart_2026_05_12.pdf",""),
    (1,"Control Environment","1.20","Goals/objectives communicated to staff","tbd","tbd","","","Confirm with Audit Manager"),
    (1,"Control Environment","1.21","New hire checklist","remote","remote","","","Pulled from 1.2 sample"),
    (1,"Control Environment","1.22","Annual performance review listing (employee name, supervisor, review date since June 1 2025)","provide","collected","","Deliverables\\1.2,1.4,1.22 Staffing and Performance Reviews.xlsx",""),
    (1,"Control Environment","1.23","Termination checklist","remote","remote","","","Pulled from 1.2 sample"),

    # Matrix 2 — Physical Security
    (2,"Physical Security","2.1","Proximity card / physical access to corporate suite","remote","remote","","",""),
    (2,"Physical Security","2.2","Badge removal on termination","remote","remote","","",""),
    (2,"Physical Security","2.3","Locked filing cabinets for sensitive documents","remote","remote","","",""),
    (2,"Physical Security","2.4","On-site shredding","remote","remote","","",""),
    (2,"Physical Security","2.5","Microsoft Azure physical controls","remote","remote","","",""),
    (2,"Physical Security","2.6","Most recent SOC 1 Type 2 or SOC 2 Type 2 report from data center (Azure)","provide","collected","","Deliverables\\SOC Certs\\Microsoft 365 Central Services - SOC 2 Type 2 Report (9-30-2025).pdf + SOC Certs\\Microsoft 365 Microservices (Type 2) - SOC 2 Type 2 Report (9-30-2025).pdf + SOC Certs\\bridge letters","Satisfies 2.6, 2.7, 3.2, 5.7, 5.13, 8.9"),
    (2,"Physical Security","2.7","Data center SOC report (physical access controls)","provide","collected","","Same as 2.6",""),
    (2,"Physical Security","2.8","Management memo documenting review of data center SOC report","provide","collected","","Deliverables\\Management Memo - Data Center SOC Report Review.pdf","Satisfies 2.8, 3.3, 5.8, 5.14, 8.10"),

    # Matrix 3 — Environmental Security
    (3,"Environmental Security","3.1","Azure environmental controls","remote","remote","","",""),
    (3,"Environmental Security","3.2","Data center environmental SOC report","provide","collected","","Same as 2.6",""),
    (3,"Environmental Security","3.3","Management memo reviewing SOC report","provide","collected","","Same as 2.8",""),

    # Matrix 4 — Computer Operations I: Backups
    (4,"Backups","4.1","Backup policy — remote review","remote","remote","","","Datto SaaS Protection"),
    (4,"Backups","4.2","Backup logs / alerts","remote","remote","","",""),
    (4,"Backups","4.3","Backup encryption at rest","remote","remote","","",""),

    # Matrix 5 — Computer Operations II: System Uptime
    (5,"System Uptime","5.1","Annual risk assessment documentation","provide","collected","","Deliverables\\5.1 -Risk Assessment 2026.docx",""),
    (5,"System Uptime","5.2","Disaster Recovery plan + Business Continuity plan","provide","collected","","Deliverables\\5.2-Business Continuity and Disaster Recovery Plan.docx",""),
    (5,"System Uptime","5.3","Sample customer agreement","provide","collected","","Deliverables\\Master Services Agreement - Template 2025-05-12.docx + TSD Proposal Template.docx",""),
    (5,"System Uptime","5.4","Infrastructure change management policies","provide","collected","","Deliverables\\Infrastructure Change Management Policy.docx",""),
    (5,"System Uptime","5.5","Closed infrastructure change tickets since June 1 2025 (Excel/CSV: ticket #, open date, close date, requestor, performer)","provide","collected","","Deliverables\\5.5 Closed Infrastructure Alert Tickets.xlsx",""),
    (5,"System Uptime","5.6","Tickets/change forms list (same format as 5.5)","provide","collected","","Deliverables\\5.6 Closed Change Management Tickets.xlsx",""),
    (5,"System Uptime","5.7","Data center SOC report","provide","collected","","Same as 2.6",""),
    (5,"System Uptime","5.8","Management memo reviewing SOC report","provide","collected","","Same as 2.8",""),
    (5,"System Uptime","5.9","Helpdesk ticketing — remote review","remote","remote","","","HaloPSA"),
    (5,"System Uptime","5.10","SLA / policies / RTO / RPO documentation on acceptable downtime levels","provide","collected","","Deliverables\\5.10-Business Continuity Communication Guide.docx",""),
    (5,"System Uptime","5.11","Enterprise monitoring app screenshot — alert thresholds/configs","provide","collected","","Deliverables\\5.11a-NinjaOneMonitoring.png + Deliverables\\5.11b-NinjaOneMonitoring.png",""),
    (5,"System Uptime","5.12","Enterprise monitoring app screenshot — alert thresholds/configs","provide","collected","","Deliverables\\5.12-PagerDutyAlerting.png",""),
    (5,"System Uptime","5.13","Data center SOC report","provide","collected","","Same as 2.6",""),
    (5,"System Uptime","5.14","Management memo reviewing SOC report","provide","collected","","Same as 2.8",""),
    (5,"System Uptime","5.15","Scheduled maintenance windows","remote","remote","","",""),
    (5,"System Uptime","5.16","Current agreements with critical third-party vendors","provide","collected","","Deliverables\\Vendor - HaloPSA USA Contract 2024.pdf + Vendor - HubSpot Signed Data Processing Agreement.pdf + Vendor - KnowBe4 - Terms of Service.docx + Vendor - Todyl Partner Agreement and DPA.docx + Vendor - Todyl_Mutual_Confidentiality_Agreement.pdf",""),
    (5,"System Uptime","5.17","Operational methodology","remote","remote","","",""),
    (5,"System Uptime","5.18","Antivirus installed on servers","remote","remote","","",""),
    (5,"System Uptime","5.19","Antivirus app screenshot — configs/settings","provide","collected","","Deliverables\\5.19a-AV Config.png + Deliverables\\5.19b-AV Config.png + Deliverables\\5.19c-AV Config.png",""),
    (5,"System Uptime","5.20","Network-level antivirus","remote","remote","","",""),
    (5,"System Uptime","5.21","Antivirus + spam filtering","remote","remote","","",""),
    (5,"System Uptime","5.22","Workstation patch management","remote","remote","","",""),
    (5,"System Uptime","5.23","List of all workstations and laptops (names)","provide","collected","","Deliverables\\5.23a-REV_NinjaOne_Workstations_6-1-26.xlsx + Deliverables\\5.23b-Antivirus-2026-06-01 191543.pdf",""),

    # Matrix 6 — Information Security
    (6,"Information Security","6.1","Documented information security policies and procedures","provide","collected","","Deliverables\\6.1 Security-Policy.pdf",""),
    (6,"Information Security","6.2","2 recent communications from security info services (attack/vulnerability warnings)","provide","collected","","Deliverables\\6.2 GTIA Security advisory - 2026-05-11.pdf + 6.2 Todyl Security advisory - 2026-05-20.pdf",""),
    (6,"Information Security","6.3","Updated network diagram","provide","collected","","Deliverables\\6.3-Revolution Group Network Diagram 202405.pdf",""),
    (6,"Information Security","6.4","Security monitoring — remote review","remote","remote","","",""),
    (6,"Information Security","6.5","Most recent vulnerability assessment tool (VAT) test report","provide","collected","","Deliverables\\6.5-REV-MSSP-Report-202505.pdf",""),
    (6,"Information Security","6.6","Internal security assessments","remote","remote","","",""),
    (6,"Information Security","6.7","Security / change management / policies review","remote","remote","","",""),
    (6,"Information Security","6.8","Access revocation on termination","remote","remote","","",""),
    (6,"Information Security","6.9","Dates of documented user access reviews since June 1 2025 (quarterly)","provide","collected","","Deliverables\\6.9,6.10- Quarterly Security Access Review Log.docx",""),
    (6,"Information Security","6.10","Dates of documented admin access reviews since June 1 2025 (quarterly)","provide","collected","","Deliverables\\6.9,6.10- Quarterly Security Access Review Log.docx",""),
    (6,"Information Security","6.11","Software install restrictions","remote","remote","","",""),
    (6,"Information Security","6.12","Names and titles of staff with admin rights: Network / AD / DB / Applications","provide","collected","","Deliverables\\6.12 - Entra AD Priviledged Roles.png + 6.12b - Entra AD Priviledged Roles - PIM Admin.png + 6.12c - Entra AD Priviledged Roles - PIM Groups.png",""),
    (6,"Information Security","6.13","Personal data collection notice","remote","remote","","",""),
    (6,"Information Security","6.14","Encryption for sensitive data","remote","remote","","",""),
    (6,"Information Security","6.15","Sensitive data destruction","remote","remote","","",""),
    (6,"Information Security","6.16","Documented disposal of confidential info","remote","remote","","",""),
    (6,"Information Security","6.17","Data retention and destruction policies","provide","collected","","Deliverables\\6.17 - Data Destruction Policy.docx",""),
    (6,"Information Security","6.18","Third-party media destruction vendor","remote","remote","","",""),
    (6,"Information Security","6.19","List of vendors with access to sensitive data onboarded since June 1 2025 (names + NDAs)","provide","collected","","Deliverables\\6.19 Vendors Onboarded since June 1, 2025.docx",""),
    (6,"Information Security","6.20","Vendor system review","remote","remote","","Deliverables\\SOC Certs\\Halo Services Solutions Ltd. - 2025 SOC 2 Type 2 Report.pdf + SOC Certs\\Intuit QuickBooks Online Ecosystem SOC 2 Type II (09-01-2024 to 04-30-2025).pdf + SOC Certs\\HubSpot - 2025 SOC 2 Type 2 - Report.pdf + SOC Certs\\bridge letters","HaloPSA + QuickBooks + HubSpot SOC 2 reports on hand"),
    (6,"Information Security","6.21","User authentication (unique ID + password)","remote","remote","","",""),
    (6,"Information Security","6.22","Internal domain password config screenshot (complexity, length, expiry)","provide","collected","","Deliverables\\6.22 - Password Expiration Policy.png",""),
    (6,"Information Security","6.23","Account lockout config screenshot","provide","collected","","Deliverables\\6.23 - Account Lockout Settings.png",""),
    (6,"Information Security","6.24","Internal network duty segregation","remote","remote","","",""),
    (6,"Information Security","6.25","Role-based access rights","remote","remote","","",""),
    (6,"Information Security","6.26","Network event logging","remote","remote","","",""),
    (6,"Information Security","6.27","DB / app server OS account policies","remote","remote","","",""),
    (6,"Information Security","6.28","Production server authentication","remote","remote","","",""),
    (6,"Information Security","6.29","Production domain auth settings screenshot","provide","missing","","",""),
    (6,"Information Security","6.30","Names and titles of staff with specific production access","provide","collected","","Deliverables\\6.30a - PIM Entra Role - Admin Group Membership.png + 6.30b - PIM Entra Role - Support Group Membership.png + 6.30c - PIM Entra Role - Approvers Group Membership.png",""),
    (6,"Information Security","6.31","Production server event logging","remote","remote","","",""),
    (6,"Information Security","6.32","Application user authentication","remote","remote","","",""),
    (6,"Information Security","6.33","Security group access management","remote","remote","","",""),

    # Matrix 7 — Communications
    (7,"Communications","7.1","Firewall systems in place","remote","remote","","",""),
    (7,"Communications","7.2","Names and titles of staff with firewall admin rights","provide","missing","","",""),
    (7,"Communications","7.3","Firewall content filtering","remote","remote","","",""),
    (7,"Communications","7.4","Intrusion prevention system (IPS)","remote","remote","","",""),
    (7,"Communications","7.5","IPS mitigation configs","remote","remote","","",""),
    (7,"Communications","7.6","Guest wireless access point","remote","remote","","",""),
    (7,"Communications","7.7","Encrypted communication sessions","remote","remote","","",""),
    (7,"Communications","7.8","Web-based application encryption","remote","remote","","",""),
    (7,"Communications","7.9","Data access restricted to application","remote","remote","","",""),

    # Matrix 8 — Application Development
    (8,"Application Development","8.1","Documented coding standards","provide","n/a","","","Matrix 8 excluded"),
    (8,"Application Development","8.2","Dev environment segregation","remote","remote","","",""),
    (8,"Application Development","8.3","Developers have no prod access","remote","remote","","",""),
    (8,"Application Development","8.4","Version control software","remote","remote","","",""),
    (8,"Application Development","8.5","Names and titles with version control access","provide","n/a","","","Matrix 8 excluded"),
    (8,"Application Development","8.6","Change management tickets (Excel/CSV: ticket #, open, close, requestor, performer, QA, approver)","provide","n/a","","","Matrix 8 excluded"),
    (8,"Application Development","8.7","Tickets/change forms list (same format as 8.6)","provide","n/a","","","Matrix 8 excluded"),
    (8,"Application Development","8.8","Versioning history since June 1 2025","provide","n/a","","","Matrix 8 excluded"),
    (8,"Application Development","8.9","Data center SOC report","provide","n/a","","","Matrix 8 excluded"),
    (8,"Application Development","8.10","Management memo reviewing SOC report","provide","n/a","","","Matrix 8 excluded"),
    (8,"Application Development","8.11","QA testing","remote","remote","","",""),
    (8,"Application Development","8.12","Names and titles with QA environment access","provide","n/a","","","Matrix 8 excluded"),
    (8,"Application Development","8.13","Multi-phase application testing","remote","remote","","",""),
    (8,"Application Development","8.14","Business approval before production deployment","remote","remote","","",""),
]

COLUMNS = [
    ("Matrix #",        5),
    ("Matrix Name",    22),
    ("Control #",       9),
    ("Description",    52),
    ("Type",           11),
    ("Status",         13),
    ("Assigned To",    15),
    ("Evidence / File Location", 45),
    ("Notes",          35),
]

TYPE_LABELS = {
    "provide": "Provide",
    "remote":  "Remote Visit",
    "tbd":     "TBD",
}

STATUS_LABELS = {
    "collected":   "Collected ✓",
    "in-progress": "In Progress",
    "missing":     "Missing",
    "remote":      "Remote Visit",
    "tbd":         "TBD",
    "n/a":         "N/A",
}

def apply_cell(cell, value, fill=None, font=None, align_wrap=False, align_center=False, border=True):
    cell.value = value
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    else:
        cell.font = NORMAL_FONT
    if align_wrap:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    if align_center:
        cell.alignment = Alignment(horizontal="center", vertical="center")
    if border:
        cell.border = BORDER


def build_tracker():
    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------
    # Sheet 1 — Master Tracker
    # ------------------------------------------------------------------
    ws = wb.active
    ws.title = "Master Tracker"
    ws.freeze_panes = "A3"

    # Title row
    ws.row_dimensions[1].height = 28
    title_cell = ws.cell(row=1, column=1, value="Revolution Group — SOC 2026 Audit Tracker")
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = HDR_FILL
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))

    # Sub-title row
    ws.row_dimensions[2].height = 18
    sub_cell = ws.cell(row=2, column=1,
        value="Review period: June 1, 2025 – May 31, 2026  |  Information requests due: June 16, 2026  |  47 documents to provide (Matrix 8 excluded)")
    sub_cell.font = Font(italic=True, size=9, color="FFFFFF")
    sub_cell.fill = SUB_FILL
    sub_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUMNS))

    # Header row
    ws.row_dimensions[3].height = 22
    for col_idx, (header, _) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=3, column=col_idx, value=header)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

    # Column widths
    for col_idx, (_, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    data_row = 4
    current_matrix = None
    for rec in CONTROLS:
        mat_num, mat_name, ctrl, desc, typ, status, assigned, evidence, notes = rec

        # Section divider row on matrix change
        if mat_num != current_matrix:
            current_matrix = mat_num
            ws.row_dimensions[data_row].height = 16
            section_label = f"Matrix {mat_num} — {mat_name}"
            for c_idx in range(1, len(COLUMNS) + 1):
                cell = ws.cell(row=data_row, column=c_idx)
                cell.fill = SUB_FILL
                cell.border = BORDER
                cell.font = SUB_FONT
                cell.alignment = Alignment(vertical="center")
            ws.cell(row=data_row, column=1).value = section_label
            ws.merge_cells(start_row=data_row, start_column=1,
                           end_row=data_row, end_column=len(COLUMNS))
            ws.cell(row=data_row, column=1).alignment = Alignment(
                vertical="center", indent=1)
            data_row += 1

        ws.row_dimensions[data_row].height = 30
        fill = STATUS_FILL.get(status, WHITE_FILL)

        apply_cell(ws.cell(row=data_row, column=1), mat_num,
                   fill=fill, align_center=True)
        apply_cell(ws.cell(row=data_row, column=2), mat_name,
                   fill=fill, align_wrap=True)
        apply_cell(ws.cell(row=data_row, column=3), ctrl,
                   fill=fill, align_center=True)
        apply_cell(ws.cell(row=data_row, column=4), desc,
                   fill=fill, align_wrap=True)
        apply_cell(ws.cell(row=data_row, column=5),
                   TYPE_LABELS.get(typ, typ), fill=fill, align_center=True)
        apply_cell(ws.cell(row=data_row, column=6),
                   STATUS_LABELS.get(status, status), fill=fill, align_center=True)
        apply_cell(ws.cell(row=data_row, column=7), assigned,
                   fill=fill, align_wrap=True)
        apply_cell(ws.cell(row=data_row, column=8), evidence,
                   fill=fill, align_wrap=True)
        apply_cell(ws.cell(row=data_row, column=9), notes,
                   fill=fill, align_wrap=True)

        data_row += 1

    # ------------------------------------------------------------------
    # Sheet 2 — To-Do (provide items only, not collected/remote)
    # ------------------------------------------------------------------
    ws2 = wb.create_sheet("Action Items")
    ws2.freeze_panes = "A3"
    ws2.row_dimensions[1].height = 28

    t1 = ws2.cell(row=1, column=1,
                  value="Revolution Group — SOC 2026 Action Items (Provide items only)")
    t1.font = Font(bold=True, size=14, color="FFFFFF")
    t1.fill = HDR_FILL
    t1.alignment = Alignment(horizontal="left", vertical="center")
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))

    t2 = ws2.cell(row=2, column=1,
                  value="Filtered to: Type = Provide | Excludes: collected, remote, tbd rows")
    t2.font = Font(italic=True, size=9, color="FFFFFF")
    t2.fill = SUB_FILL
    t2.alignment = Alignment(horizontal="left", vertical="center")
    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUMNS))

    ws2.row_dimensions[3].height = 22
    for col_idx, (header, _) in enumerate(COLUMNS, start=1):
        c = ws2.cell(row=3, column=col_idx, value=header)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

    for col_idx, (_, width) in enumerate(COLUMNS, start=1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    action_row = 4
    for rec in CONTROLS:
        mat_num, mat_name, ctrl, desc, typ, status, assigned, evidence, notes = rec
        if typ != "provide":
            continue
        # Sort order: missing first, then in-progress, then collected
        ws2.row_dimensions[action_row].height = 30
        fill = STATUS_FILL.get(status, WHITE_FILL)

        apply_cell(ws2.cell(row=action_row, column=1), mat_num,
                   fill=fill, align_center=True)
        apply_cell(ws2.cell(row=action_row, column=2), mat_name,
                   fill=fill, align_wrap=True)
        apply_cell(ws2.cell(row=action_row, column=3), ctrl,
                   fill=fill, align_center=True)
        apply_cell(ws2.cell(row=action_row, column=4), desc,
                   fill=fill, align_wrap=True)
        apply_cell(ws2.cell(row=action_row, column=5),
                   TYPE_LABELS.get(typ, typ), fill=fill, align_center=True)
        apply_cell(ws2.cell(row=action_row, column=6),
                   STATUS_LABELS.get(status, status), fill=fill, align_center=True)
        apply_cell(ws2.cell(row=action_row, column=7), assigned,
                   fill=fill, align_wrap=True)
        apply_cell(ws2.cell(row=action_row, column=8), evidence,
                   fill=fill, align_wrap=True)
        apply_cell(ws2.cell(row=action_row, column=9), notes,
                   fill=fill, align_wrap=True)

        action_row += 1

    # ------------------------------------------------------------------
    # Sheet 3 — Summary
    # ------------------------------------------------------------------
    ws3 = wb.create_sheet("Summary")
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 8
    ws3.column_dimensions["C"].width = 8
    ws3.column_dimensions["D"].width = 8
    ws3.column_dimensions["E"].width = 8
    ws3.column_dimensions["F"].width = 8
    ws3.column_dimensions["G"].width = 12

    # Title
    ws3.row_dimensions[1].height = 28
    tc = ws3.cell(row=1, column=1, value="SOC 2026 — Summary by Matrix")
    tc.font = Font(bold=True, size=14, color="FFFFFF")
    tc.fill = HDR_FILL
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws3.merge_cells("A1:G1")

    # Header
    ws3.row_dimensions[2].height = 20
    summary_headers = ["Matrix", "Total", "Provide", "Remote", "TBD", "Collected", "% Complete"]
    for ci, h in enumerate(summary_headers, 1):
        c = ws3.cell(row=2, column=ci, value=h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

    matrix_meta = [
        (1, "Control Environment"),
        (2, "Physical Security"),
        (3, "Environmental Security"),
        (4, "Backups"),
        (5, "System Uptime"),
        (6, "Information Security"),
        (7, "Communications"),
        (8, "Application Development"),
    ]

    sr = 3
    totals = [0, 0, 0, 0, 0]
    for mnum, mname in matrix_meta:
        recs = [r for r in CONTROLS if r[0] == mnum]
        total     = len(recs)
        provide   = sum(1 for r in recs if r[4] == "provide")
        remote    = sum(1 for r in recs if r[4] == "remote")
        tbd       = sum(1 for r in recs if r[4] == "tbd")
        collected = sum(1 for r in recs if r[5] == "collected")
        pct = f"{round(collected/provide*100)}%" if provide else "N/A"
        row_vals = [f"Matrix {mnum} — {mname}", total, provide, remote, tbd, collected, pct]
        for ci, v in enumerate(row_vals, 1):
            c = ws3.cell(row=sr, column=ci, value=v)
            c.border = BORDER
            c.font = NORMAL_FONT
            c.alignment = Alignment(horizontal="center" if ci > 1 else "left",
                                    vertical="center")
        totals[0] += total; totals[1] += provide; totals[2] += remote
        totals[3] += tbd; totals[4] += collected
        sr += 1

    # Total row
    ws3.row_dimensions[sr].height = 18
    total_pct = f"{round(totals[4]/totals[1]*100)}%" if totals[1] else "N/A"
    total_vals = ["TOTAL", totals[0], totals[1], totals[2], totals[3], totals[4], total_pct]
    for ci, v in enumerate(total_vals, 1):
        c = ws3.cell(row=sr, column=ci, value=v)
        c.border = BORDER
        c.font = Font(bold=True, size=10)
        c.fill = GREY_FILL
        c.alignment = Alignment(horizontal="center" if ci > 1 else "left",
                                vertical="center")

    sr += 2

    # Legend
    ws3.cell(row=sr, column=1, value="Colour Legend").font = Font(bold=True, size=10)
    sr += 1
    legend = [
        ("Missing", RED_FILL),
        ("In Progress", YELLOW_FILL),
        ("Collected ✓", GREEN_FILL),
        ("Remote Visit (July)", BLUE_FILL),
        ("TBD / N/A", GREY_FILL),
    ]
    for label, fill in legend:
        c = ws3.cell(row=sr, column=1, value=label)
        c.fill = fill
        c.font = NORMAL_FONT
        c.border = BORDER
        sr += 1

    wb.save(WB_PATH)
    print(f"Saved: {WB_PATH}")


if __name__ == "__main__":
    build_tracker()
